from io import BytesIO
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import Response as HTTPResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from database import get_db
from models import Product
from auth import get_current_user, require_write

router = APIRouter(prefix="/products", tags=["products"])


class ProductIn(BaseModel):
    type: str
    room: str
    description: str
    details: str | None = None
    uom: str | None = None
    price: float = 0.0
    is_active: bool = True


def _dict(p: Product) -> dict:
    return {
        "id": p.id, "type": p.type, "room": p.room, "description": p.description,
        "details": p.details, "uom": p.uom, "price": p.price, "is_active": p.is_active,
    }


@router.get("")
def list_products(db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    return [_dict(p) for p in db.query(Product).filter(Product.is_active == True).order_by(Product.type, Product.room, Product.description).all()]


@router.get("/all")
def list_all_products(db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    return [_dict(p) for p in db.query(Product).order_by(Product.type, Product.room, Product.description).all()]


@router.post("", status_code=201)
def create_product(data: ProductIn, db: Session = Depends(get_db), user: dict = Depends(require_write)):
    p = Product(**data.model_dump())
    db.add(p)
    db.commit()
    db.refresh(p)
    return _dict(p)


@router.patch("/{product_id}")
def update_product(product_id: int, data: ProductIn, db: Session = Depends(get_db), user: dict = Depends(require_write)):
    p = db.get(Product, product_id)
    if not p:
        raise HTTPException(404, "Product not found")
    for k, v in data.model_dump().items():
        setattr(p, k, v)
    p.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(p)
    return _dict(p)


@router.delete("/{product_id}", status_code=204)
def delete_product(product_id: int, db: Session = Depends(get_db), user: dict = Depends(require_write)):
    p = db.get(Product, product_id)
    if not p:
        raise HTTPException(404, "Product not found")
    db.delete(p)
    db.commit()


COLS = ["Type", "Room", "Description", "Details", "UOM", "Price", "Active"]
LEGACY_COLS = ["Type", "Room", "Description", "UOM", "Price", "Active"]  # pre-Details files


@router.get("/export-xlsx")
def export_xlsx(db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    for c, h in enumerate(COLS, start=1):
        cell = ws.cell(1, c, h)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="6D28D9")

    r = 2
    for p in db.query(Product).order_by(Product.type, Product.room, Product.description).all():
        ws.cell(r, 1, p.type)
        ws.cell(r, 2, p.room)
        ws.cell(r, 3, p.description)
        ws.cell(r, 4, p.details)
        ws.cell(r, 5, p.uom)
        ws.cell(r, 6, p.price)
        ws.cell(r, 7, "Y" if p.is_active else "")
        r += 1

    if r == 2:  # empty table → template row
        ws.cell(2, 1, "MFC"); ws.cell(2, 2, "Kitchen")
        ws.cell(2, 3, "MFC Wall Cabinet (300mm D x 700mm H)")
        ws.cell(2, 4, "Blum soft-close hinges · matte laminate finish")
        ws.cell(2, 5, "Fr"); ws.cell(2, 6, 330); ws.cell(2, 7, "Y")

    widths = [20, 16, 52, 44, 10, 12, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return HTTPResponse(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="products.xlsx"'},
    )


@router.post("/import-xlsx")
async def import_xlsx(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(require_write),
):
    from openpyxl import load_workbook
    try:
        wb = load_workbook(BytesIO(await file.read()), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Not a valid xlsx: {e}")
    ws = wb.active

    # Accept both the new 7-column header (with Details) and the legacy 6-column
    # header (without Details), so pre-existing product spreadsheets still import.
    new_header = [(str(ws.cell(1, c).value).strip() if ws.cell(1, c).value else "") for c in range(1, len(COLS) + 1)]
    legacy_header = [(str(ws.cell(1, c).value).strip() if ws.cell(1, c).value else "") for c in range(1, len(LEGACY_COLS) + 1)]
    if new_header == COLS:
        has_details = True
    elif legacy_header == LEGACY_COLS:
        has_details = False
    else:
        raise HTTPException(400, f"Header row must be: {', '.join(COLS)}")

    parsed = []
    for r in range(2, ws.max_row + 1):
        if has_details:
            typ, room, desc, details, uom, price, active = [ws.cell(r, c).value for c in range(1, 8)]
        else:
            typ, room, desc, uom, price, active = [ws.cell(r, c).value for c in range(1, 7)]
            details = None
        if not any([typ, room, desc]):
            continue  # blank row
        if not desc:
            raise HTTPException(400, f"Row {r}: Description is required")
        try:
            price_val = float(price) if price not in (None, "") else 0.0
        except (ValueError, TypeError):
            raise HTTPException(400, f"Row {r}: Price '{price}' is not a number")
        parsed.append({
            "type": str(typ).strip() if typ else "",
            "room": str(room).strip() if room else "",
            "description": str(desc).strip(),
            "details": (str(details).strip() if details else None),
            "uom": (str(uom).strip() if uom else None),
            "price": price_val,
            "is_active": str(active).strip().upper() in ("Y", "YES", "TRUE", "1") if active not in (None, "") else True,
        })

    db.query(Product).delete()
    for row in parsed:
        db.add(Product(**row))
    db.commit()
    return {"ok": True, "products": len(parsed)}
