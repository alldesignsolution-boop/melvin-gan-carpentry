"""Preset library — DB-backed hierarchical quotation presets.

Tree shape: Area -> SubArea -> Set -> Items (matches the shape the
quote_form.html JS expects).
"""
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import Response as HTTPResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from database import get_db
from models import PresetArea, PresetSubArea, PresetSet, PresetItem
from auth import get_current_user, require_write

router = APIRouter(prefix="/presets", tags=["presets"])


def _tree(db: Session) -> list[dict]:
    areas = db.query(PresetArea).order_by(PresetArea.sort_order, PresetArea.id).all()
    return [
        {
            "area": a.name,
            "targetRoom": a.target_room,
            "subs": [
                {
                    "name": sub.name,
                    "sets": [
                        {
                            "name": s.name,
                            "runFt": s.run_ft,
                            "category": s.category,
                            "items": [
                                {
                                    "description": it.description,
                                    "uom": it.uom,
                                    "isPerFt": it.is_per_ft,
                                    "qty": it.qty,
                                    "unit_price": it.unit_price,
                                }
                                for it in s.items
                            ],
                        }
                        for s in sub.sets
                    ],
                }
                for sub in a.subs
            ],
        }
        for a in areas
    ]


@router.get("")
def get_presets(db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    return _tree(db)


# ── Full-tree replace (used by the inline editor on the Library page) ──────────
# The editor sends a flattened Category → Group → Items structure; we persist it
# using the existing schema with one implicit Set per group.
class PItemIn(BaseModel):
    description: str
    uom: str | None = ""
    isPerFt: bool = False
    qty: float = 1.0
    unit_price: float = 0.0


class PSetIn(BaseModel):
    name: str = ""
    runFt: float | None = None
    category: str = "construction"
    items: list[PItemIn] = []


class PSubIn(BaseModel):
    name: str
    sets: list[PSetIn] = []


class PAreaIn(BaseModel):
    area: str
    targetRoom: str | None = None
    subs: list[PSubIn] = []


@router.put("")
def replace_presets(areas: list[PAreaIn], db: Session = Depends(get_db),
                    user: dict = Depends(require_write)):
    db.query(PresetItem).delete()
    db.query(PresetSet).delete()
    db.query(PresetSubArea).delete()
    db.query(PresetArea).delete()
    db.flush()

    n_sets = n_items = 0
    for ai, a in enumerate(areas):
        area = PresetArea(
            name=(a.area or "").strip() or "Category",
            target_room=((a.targetRoom or a.area) or "").strip() or "Category",
            sort_order=ai,
        )
        db.add(area); db.flush()
        for si, sub in enumerate(a.subs):
            subrow = PresetSubArea(area_id=area.id, name=(sub.name or "").strip() or "Group", sort_order=si)
            db.add(subrow); db.flush()
            for sti, st in enumerate(sub.sets):
                setrow = PresetSet(
                    sub_area_id=subrow.id,
                    name=(st.name or sub.name or "Set").strip(),
                    run_ft=st.runFt,
                    category=(st.category or "construction"),
                    sort_order=sti,
                )
                db.add(setrow); db.flush()
                n_sets += 1
                for iti, it in enumerate(st.items):
                    desc = (it.description or "").strip()
                    if not desc:
                        continue
                    db.add(PresetItem(
                        set_id=setrow.id, description=desc, uom=(it.uom or ""),
                        is_per_ft=bool(it.isPerFt), qty=it.qty, unit_price=it.unit_price,
                        sort_order=iti,
                    ))
                    n_items += 1
    db.commit()
    return {"ok": True, "areas": len(areas), "sets": n_sets, "items": n_items}


COLS = ["Area", "TargetRoom", "SubArea", "Set", "RunFt", "Category",
        "Description", "UOM", "IsPerFt", "Qty", "UnitPrice"]


@router.get("/export-xlsx")
def export_xlsx(db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Presets"

    # Header row
    for c, h in enumerate(COLS, start=1):
        cell = ws.cell(1, c, h)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="6D28D9")

    r = 2
    areas = db.query(PresetArea).order_by(PresetArea.sort_order, PresetArea.id).all()
    for a in areas:
        for sub in a.subs:
            for s in sub.sets:
                for it in s.items:
                    ws.cell(r, 1, a.name)
                    ws.cell(r, 2, a.target_room)
                    ws.cell(r, 3, sub.name)
                    ws.cell(r, 4, s.name)
                    ws.cell(r, 5, s.run_ft)
                    ws.cell(r, 6, s.category)
                    ws.cell(r, 7, it.description)
                    ws.cell(r, 8, it.uom)
                    ws.cell(r, 9, "Y" if it.is_per_ft else "")
                    ws.cell(r, 10, it.qty)
                    ws.cell(r, 11, it.unit_price)
                    r += 1

    # If DB is empty, add a template row so user has a starting point
    if r == 2:
        ws.cell(2, 1, "Kitchen"); ws.cell(2, 2, "Kitchen")
        ws.cell(2, 3, "Wet Area"); ws.cell(2, 4, "Default Set"); ws.cell(2, 5, 10)
        ws.cell(2, 6, "construction")
        ws.cell(2, 7, "Wall Cabinet 300D × 700H (MFC)"); ws.cell(2, 8, "ft")
        ws.cell(2, 9, "Y"); ws.cell(2, 10, 1); ws.cell(2, 11, 330)

    widths = [16, 18, 18, 26, 10, 14, 44, 8, 10, 8, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return HTTPResponse(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="presets.xlsx"'},
    )


@router.post("/import-xlsx")
async def import_xlsx(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(require_write),
):
    from openpyxl import load_workbook
    try:
        wb = load_workbook(BytesIO(await file.read()), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Not a valid xlsx: {e}")
    ws = wb.active

    # Validate header
    header = [(ws.cell(1, c).value or "").strip() for c in range(1, len(COLS) + 1)]
    if header != COLS:
        raise HTTPException(400, f"Header row must be: {', '.join(COLS)}")

    # Parse rows
    parsed = []  # list of dicts
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, len(COLS) + 1)]
        area, target_room, sub, set_name, run_ft, category, desc, uom, is_perft, qty, unit_price = row
        if not any([area, sub, set_name, desc]):
            continue  # blank
        if not (area and sub and set_name and desc):
            raise HTTPException(400, f"Row {r}: missing required field (Area/SubArea/Set/Description)")
        parsed.append({
            "area": str(area).strip(),
            "target_room": str(target_room or area).strip(),
            "sub": str(sub).strip(),
            "set": str(set_name).strip(),
            "run_ft": float(run_ft) if run_ft not in (None, "") else None,
            "category": (str(category).strip().lower() if category else "construction"),
            "description": str(desc).strip(),
            "uom": (str(uom).strip() if uom else ""),
            "is_per_ft": str(is_perft).strip().upper() in ("Y", "YES", "TRUE", "1"),
            "qty": float(qty) if qty not in (None, "") else 1.0,
            "unit_price": float(unit_price) if unit_price not in (None, "") else 0.0,
        })

    # Wipe and rebuild
    db.query(PresetItem).delete()
    db.query(PresetSet).delete()
    db.query(PresetSubArea).delete()
    db.query(PresetArea).delete()
    db.flush()

    areas: dict[str, PresetArea] = {}
    subs: dict[tuple[str, str], PresetSubArea] = {}
    sets: dict[tuple[str, str, str], PresetSet] = {}

    for i, row in enumerate(parsed):
        a_key = row["area"]
        if a_key not in areas:
            a = PresetArea(name=row["area"], target_room=row["target_room"], sort_order=len(areas))
            db.add(a); db.flush()
            areas[a_key] = a
        area_obj = areas[a_key]

        s_key = (a_key, row["sub"])
        if s_key not in subs:
            sub_obj = PresetSubArea(area_id=area_obj.id, name=row["sub"], sort_order=len(subs))
            db.add(sub_obj); db.flush()
            subs[s_key] = sub_obj
        sub_obj = subs[s_key]

        set_key = (a_key, row["sub"], row["set"])
        if set_key not in sets:
            set_obj = PresetSet(
                sub_area_id=sub_obj.id, name=row["set"],
                run_ft=row["run_ft"], category=row["category"], sort_order=len(sets),
            )
            db.add(set_obj); db.flush()
            sets[set_key] = set_obj
        set_obj = sets[set_key]

        item = PresetItem(
            set_id=set_obj.id, description=row["description"], uom=row["uom"],
            is_per_ft=row["is_per_ft"], qty=row["qty"], unit_price=row["unit_price"],
            sort_order=len([x for x in parsed[:i] if (x["area"], x["sub"], x["set"]) == set_key]),
        )
        db.add(item)

    db.commit()
    return {"ok": True, "areas": len(areas), "subs": len(subs), "sets": len(sets), "items": len(parsed)}
